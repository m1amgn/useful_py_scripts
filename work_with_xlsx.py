from openpyxl import Workbook, load_workbook

## loading previous workbook
wb = load_workbook('SampleData.xlsx')

## access the content from the active sheet
ws = wb.active
print(ws)


## Accessing other sheets
ws = wb["SalesOrders"]
print(ws)

## Access Cell
cell_to_access= ws['A5']
print(cell_to_access)


## Accessing Cell Value
cell_value = ws['B5'].value
print(cell_value)

## Changing the value of a cell
print(ws['B1'].value)

ws['B1'].value="Location"
print(ws['B1'].value)

## Printing Sheet Names
print(wb.sheetnames)

## Creating New Sheet
wb.create_sheet("Percentage")
print(wb.sheetnames)

## Saving the sheet
wb.save("test1.xlsx")

## Create new workbook
wb = Workbook()
ws = wb.active
ws.title = "Test"

## Adding data to the worksheet
ws.append(["S.N.","A", "B", "C"])

## Adding Rows
for i in range(2):
    ws.insert_rows(5) # row_number as input

## Deleting Rows
for i in range(3):
   ws.delete_rows(3)  # column_number as the input.

## Adding new columns
ws.insert_cols(3) # column_number as the input

## Deleting columns
ws.delete_cols(2) # column_number as the input

## Moving Data
ws.move_range("B2:D9", rows=0, cols=2 )

wb.save("New_WB.xlsx")
